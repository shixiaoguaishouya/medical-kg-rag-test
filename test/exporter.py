"""Excel 导出模块：生成测试结果 Excel 文件"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=8, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val[:100])
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def export_results(results, stats, output_path):
    """导出测试结果到 Excel"""
    wb = Workbook()

    # ==================== Sheet 1: 逐题明细 ====================
    ws1 = wb.active
    ws1.title = "逐题明细"

    headers1 = [
        "题号", "课程", "分类", "难度",
        "题目", "标准答案",
        "KG-RAG回答", "RAG提取", "RAG结果",
        "LLM回答", "LLM提取", "LLM结果",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(headers1))

    from judge import is_correct

    for i, r in enumerate(results, start=2):
        rag_correct, rag_opt = is_correct(r.get("rag_answer", ""), r.get("answer", ""))
        llm_correct, llm_opt = is_correct(r.get("llm_answer", ""), r.get("answer", ""))

        row_data = [
            r.get("id", ""), r.get("course", ""), r.get("subtype", ""), r.get("difficulty", ""),
            r.get("question", ""), r.get("answer", ""),
            r.get("rag_answer", ""), rag_opt,
            "正确" if rag_correct else "错误",
            r.get("llm_answer", ""), llm_opt,
            "正确" if llm_correct else "错误",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP if col in (5, 7, 10) else CENTER
        # 着色
        for col in (9, 12):
            cell = ws1.cell(row=i, column=col)
            if cell.value == "正确":
                cell.font = Font(color="008000")
            else:
                cell.font = Font(color="CC0000")

    _auto_width(ws1)
    ws1.column_dimensions["E"].width = 45
    ws1.column_dimensions["G"].width = 50
    ws1.column_dimensions["J"].width = 50

    # ==================== Sheet 2: 按课程汇总 ====================
    ws2 = wb.create_sheet("按课程汇总")

    headers2 = ["课程", "题目数", "RAG正确", "RAG正确率", "LLM正确", "LLM正确率"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))

    course_stats = stats.get("by_course", [])
    for i, cs in enumerate(course_stats, start=2):
        row_data = [cs["course"], cs["total"], cs["rag_correct"], cs["rag_accuracy"], cs["llm_correct"], cs["llm_accuracy"]]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col in (4, 6):
                cell.number_format = "0.00%"

    total_row = len(course_stats) + 2
    overall = stats.get("overall", {})
    for col, val in enumerate(["合计", overall.get("total", 0), overall.get("rag_correct", 0),
                                overall.get("rag_accuracy", 0), overall.get("llm_correct", 0),
                                overall.get("llm_accuracy", 0)], 1):
        cell = ws2.cell(row=total_row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.font = Font(bold=True)
        if col in (4, 6):
            cell.number_format = "0.00%"

    _auto_width(ws2)

    # ==================== Sheet 3: 按难度汇总 ====================
    ws3 = wb.create_sheet("按难度汇总")

    headers3 = ["难度", "题目数", "RAG正确", "RAG正确率", "LLM正确", "LLM正确率"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers3))

    diff_stats = stats.get("by_difficulty", [])
    for i, ds in enumerate(diff_stats, start=2):
        row_data = [ds["difficulty"], ds["total"], ds["rag_correct"], ds["rag_accuracy"], ds["llm_correct"], ds["llm_accuracy"]]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col in (4, 6):
                cell.number_format = "0.00%"

    _auto_width(ws3)

    # ==================== Sheet 4: 课程x难度交叉汇总 ====================
    ws4 = wb.create_sheet("课程x难度交叉")

    # 收集所有课程和难度
    from collections import defaultdict
    courses_set = set()
    diffs_set = set()
    cross = defaultdict(lambda: defaultdict(lambda: {"total": 0, "rag_correct": 0, "llm_correct": 0}))
    for r in results:
        c = r.get("course", "未知")
        d = r.get("difficulty", "未知")
        courses_set.add(c)
        diffs_set.add(d)
        cross[c][d]["total"] += 1
        rag_ok, _ = is_correct(r.get("rag_answer", ""), r.get("answer", ""))
        llm_ok, _ = is_correct(r.get("llm_answer", ""), r.get("answer", ""))
        if rag_ok:
            cross[c][d]["rag_correct"] += 1
        if llm_ok:
            cross[c][d]["llm_correct"] += 1

    courses = sorted(courses_set)
    diffs = sorted(diffs_set)

    # 表头：课程 | 难度1-RAG正确数/率 | 难度1-LLM正确数/率 | ...
    headers4 = ["课程"]
    for d in diffs:
        headers4.append(f"{d}-题目数")
        headers4.append(f"{d}-RAG正确")
        headers4.append(f"{d}-RAG正确率")
        headers4.append(f"{d}-LLM正确")
        headers4.append(f"{d}-LLM正确率")

    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header(ws4, 1, len(headers4))

    for i, course in enumerate(courses, start=2):
        row_data = [course]
        for d in diffs:
            data = cross[course][d]
            t = data["total"]
            row_data.append(t)
            row_data.append(data["rag_correct"])
            row_data.append(round(data["rag_correct"] / t, 4) if t > 0 else 0)
            row_data.append(data["llm_correct"])
            row_data.append(round(data["llm_correct"] / t, 4) if t > 0 else 0)
        for col, val in enumerate(row_data, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col > 1 and (col % 5) in (3, 5):  # 正确率列
                cell.number_format = "0.00%"

    _auto_width(ws4)

    # ==================== Sheet 5: 总体统计 ====================
    ws5 = wb.create_sheet("总体统计")

    info_data = [
        ("测试时间", stats.get("test_time", "")),
        ("总题数", overall.get("total", 0)),
        ("", ""),
        ("KG-RAG 正确数", overall.get("rag_correct", 0)),
        ("KG-RAG 正确率", overall.get("rag_accuracy", 0)),
        ("KG-RAG 接口失败", stats.get("rag_fail_count", 0)),
        ("", ""),
        ("通用LLM 正确数", overall.get("llm_correct", 0)),
        ("通用LLM 正确率", overall.get("llm_accuracy", 0)),
        ("通用LLM 接口失败", stats.get("llm_fail_count", 0)),
        ("", ""),
        ("总耗时(秒)", stats.get("total_time_s", 0)),
    ]

    for col, h in enumerate(["指标", "数值"], 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, (label, val) in enumerate(info_data, start=2):
        ws5.cell(row=i, column=1, value=label).border = THIN_BORDER
        cell = ws5.cell(row=i, column=2, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if isinstance(val, float) and 0 < val < 1:
            cell.number_format = "0.00%"

    _auto_width(ws5)

    wb.save(output_path)
    return output_path
